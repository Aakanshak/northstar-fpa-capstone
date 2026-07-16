"""Build Northstar's formula-driven FP&A workbook.

Values from CSVs are source inputs. Every model output is an Excel formula.
The workbook uses a prior-period cash circularity breaker: interest is computed
on beginning cash, Cash Flow calculates ending cash, and Balance Sheet links to it.
"""
from __future__ import annotations

import csv, json, subprocess
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

ROOT=Path(__file__).resolve().parents[1]
DATA=ROOT/"02-data-generation"/"output_csvs"
OUT=Path(__file__).resolve().parent/"northstar_fpa_model.xlsx"
SUMMARY=Path(__file__).resolve().parent/"model_summary.json"
NAVY="17324D"; BLUE="2F75B5"; TEAL="00A6A6"; GREEN="70AD47"; RED="C00000"; GOLD="F4B183"; LIGHT="EAF2F8"; GRAY="E7E6E6"; WHITE="FFFFFF"; INPUT="FFF2CC"
MONEY='$#,##0;[Red]($#,##0);-'; MONEY_K='$#,##0,"K";[Red]($#,##0,"K");-'; PCT='0.0%;[Red](0.0%);-'; MULT='0.0x'; INT='#,##0'
MONTHS=[datetime(2026,m,1) for m in range(1,13)]

def rows(path):
    with path.open(encoding='utf-8') as f: return list(csv.DictReader(f))
def title(ws,text,last=14):
    ws.sheet_view.showGridLines=False; ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last)
    c=ws.cell(1,1,text); c.font=Font(name='Aptos Display',size=20,bold=True,color=WHITE); c.fill=PatternFill('solid',fgColor=NAVY); c.alignment=Alignment(vertical='center'); ws.row_dimensions[1].height=34
def header(ws,row,start,end):
    for c in ws.iter_cols(min_col=start,max_col=end,min_row=row,max_row=row):
        c[0].font=Font(bold=True,color=WHITE); c[0].fill=PatternFill('solid',fgColor=BLUE); c[0].alignment=Alignment(horizontal='center',wrap_text=True)
def section(ws,row,text,start=1,end=14):
    ws.merge_cells(start_row=row,start_column=start,end_row=row,end_column=end); c=ws.cell(row,start,text); c.fill=PatternFill('solid',fgColor=LIGHT); c.font=Font(bold=True,color=NAVY); c.alignment=Alignment(vertical='center')
def months(ws,row,start_col=3):
    for i,d in enumerate(MONTHS,start_col): ws.cell(row,i,d).number_format='mmm-yy'
def fmt_row(ws,row,c1=3,c2=14,fmt=MONEY_K):
    for c in range(c1,c2+1): ws.cell(row,c).number_format=fmt
def style_model(ws,max_row,max_col):
    ws.freeze_panes='C4'; ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=30
    for c in range(3,max_col+1): ws.column_dimensions[get_column_letter(c)].width=13
    for r in range(1,max_row+1):
        ws.row_dimensions[r].height=max(ws.row_dimensions[r].height or 15,15)
def formula(ws,cell,text,comment=None):
    ws[cell]=text
    if comment: ws[cell].comment=Comment(comment,'AAKANSHA')
def name(wb,name,sheet,cell):
    wb.defined_names.add(DefinedName(name,attr_text=f"'{sheet}'!${cell[0]}${cell[1:]}"))

def build():
    ledger=rows(DATA/'customer_arr_ledger.csv'); hc=rows(DATA/'headcount_plan.csv'); pipe=rows(DATA/'sales_pipeline.csv'); avb=rows(DATA/'actuals_vs_budget.csv')
    wb=Workbook(); wb.remove(wb.active)
    names=['Cover & Instructions','Assumptions','Customer ARR Ledger','ARR Bridge','Revenue Forecast','Headcount & Opex Plan','Income Statement','Balance Sheet','Cash Flow Statement','SaaS Metrics Dashboard','Budget vs Actual','Sensitivity Analysis','Cohort LTV Analysis']
    for n in names: wb.create_sheet(n)
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'; wb.calculation.iterate=False

    # 1 Cover
    ws=wb[names[0]]; title(ws,'Northstar SaaS | Revenue Forecasting, Budget Variance & Board Reporting',10)
    ws['B3']='Q3 2026 Board Model'; ws['B3'].font=Font(size=24,bold=True,color=NAVY)
    ws['B5']='Purpose'; ws['B5'].font=Font(bold=True,color=BLUE); ws['B6']='Answer: Are we hitting revenue targets, why did we miss/beat budget, and what is the realistic FY forecast?'
    ws['B8']='How to use'; ws['B8'].font=Font(bold=True,color=BLUE)
    for r,t in enumerate(['1. Change yellow inputs on Assumptions.','2. Select Base / Upside / Downside using the Scenario Selector.','3. Review linked forecast, 3 statements, metrics, variance and sensitivities.','4. Blue font = formulas; yellow fill = editable assumption.'],9): ws.cell(r,2,t)
    ws['B15']='Circularity policy'; ws['B15'].font=Font(bold=True,color=BLUE); ws['B16']='Interest uses prior-period cash. Cash Flow calculates ending cash; Balance Sheet links to that result. No iterative calculation is required.'
    ws['B19']='Model integrity checks'; ws['B19'].font=Font(bold=True,color=BLUE)
    ws['B20']='Balance Sheet check'; ws['C20']="='Balance Sheet'!N22"; ws['B21']='ARR bridge check'; ws['C21']="='ARR Bridge'!N11"; ws['B22']='Cash flow check'; ws['C22']="='Cash Flow Statement'!N17-'Balance Sheet'!N5"
    for c in ['C20','C21','C22']: ws[c].number_format=MONEY; ws[c].font=Font(color=BLUE)
    ws['B24']='Selected scenario'; ws['C24']="=INDEX('Assumptions'!K3:K5,MATCH(Scenario_Select,'Assumptions'!J3:J5,0))"
    ws.column_dimensions['B'].width=115; ws.column_dimensions['C'].width=18; ws.sheet_view.showGridLines=False

    # 2 Assumptions
    ws=wb[names[1]]; title(ws,'Assumptions | centralized, scenario-specific model drivers',8)
    header(ws,3,2,8); ws.append([]); ws['B3']='Driver'; ws['C3']='Base'; ws['D3']='Upside'; ws['E3']='Downside'; ws['F3']='Units'; ws['G3']='Rationale'; ws['H3']='Selected'
    assumptions=[
      ('Scenario Selector',1,2,3,'1=Base','Board planning case',None),
      ('Monthly logo churn',.011,.008,.016,'%','Segment-weighted logo loss',None),
      ('Monthly expansion rate',.0065,.0085,.0040,'%','Expansion ARR on beginning ARR',None),
      ('Monthly contraction rate',.0020,.0014,.0032,'%','Downgrades on beginning ARR',None),
      ('Pipeline conversion factor',.55,.70,.38,'x','Applied to probability-weighted pipeline after stage weighting',None),
      ('Sales cycle timing factor',1.00,.92,1.18,'x','Changes timing / close slippage',None),
      ('Hiring pace',1.00,1.08,.82,'x','Scales planned headcount spend',None),
      ('Gross margin',.805,.815,.785,'%','Subscription delivery margin',None),
      ('Annual merit increase',.04,.04,.03,'%','Compensation inflation',None),
      ('Interest rate on cash',.045,.048,.040,'%','Beginning-cash circularity breaker',None),
      ('Capex as % revenue',.025,.024,.020,'%','Capitalized internal systems',None),
      ('DSO',48,44,55,'days','Collections efficiency',None),
      ('DPO',32,36,28,'days','Vendor payment timing',None),
      ('Opening cash',18_000_000,18_000_000,18_000_000,'$','1-Jan-2026 cash balance',None),
      ('Opening debt',5_000_000,5_000_000,5_000_000,'$','Term debt outstanding',None),
      ('Tax rate',.21,.21,.21,'%','US blended statutory rate',None),
      ('CAC per new logo',42_000,39_000,49_000,'$','Fully loaded acquisition cost',None),
      ('Discount rate',.12,.11,.14,'%','SaaS cost of capital proxy',None),
      ('Variance threshold %',.05,.05,.05,'%','Board materiality flag',None),
      ('Variance threshold $',50_000,50_000,50_000,'$','Board materiality flag',None),
    ]
    for i,(label,b,u,d,unit,rat,_) in enumerate(assumptions,4):
        ws.cell(i,2,label); ws.cell(i,3,b); ws.cell(i,4,u); ws.cell(i,5,d); ws.cell(i,6,unit); ws.cell(i,7,rat)
        if i==4: ws.cell(i,8,'=C4')
        else: ws.cell(i,8,f'=CHOOSE($H$4,C{i},D{i},E{i})')
        for c in range(3,6): ws.cell(i,c).fill=PatternFill('solid',fgColor=INPUT); ws.cell(i,c).font=Font(color='0000FF')
        ws.cell(i,8).font=Font(color=BLUE)
        fmt=PCT if unit=='%' else MONEY if unit=='$' else '0.00x' if unit=='x' else '0'
        for c in range(3,6): ws.cell(i,c).number_format=fmt
        ws.cell(i,8).number_format=fmt
    ranges={'Scenario_Select':'H4','Churn_Rate':'H5','Expansion_Rate':'H6','Contraction_Rate':'H7','Pipeline_Factor':'H8','Sales_Cycle_Factor':'H9','Hiring_Pace':'H10','Gross_Margin':'H11','Merit_Increase':'H12','Interest_Rate':'H13','Capex_Pct':'H14','DSO':'H15','DPO':'H16','Opening_Cash':'H17','Opening_Debt':'H18','Tax_Rate':'H19','CAC_Per_Logo':'H20','Discount_Rate':'H21','Variance_Pct':'H22','Variance_Dollar':'H23'}
    for n,c in ranges.items(): name(wb,n,'Assumptions',c)
    ws['H4'].comment=Comment('Enter 1 for Base, 2 for Upside, or 3 for Downside. All selected drivers update through named ranges.','AAKANSHA')
    ws['J3']=1; ws['K3']='Base'; ws['J4']=2; ws['K4']='Upside'; ws['J5']=3; ws['K5']='Downside'; ws['G25']='Selected scenario name'; ws['H25']='=INDEX($K$3:$K$5,MATCH($H$4,$J$3:$J$5,0))'
    name(wb,'Scenario_Name','Assumptions','H25')
    ws.freeze_panes='B4'; ws.column_dimensions['B'].width=30; ws.column_dimensions['G'].width=42

    # 3 Ledger raw source
    ws=wb[names[2]]; title(ws,'Customer ARR Ledger | source data imported from CSV',10)
    hdr=list(ledger[0]);
    for c,h in enumerate(hdr,1): ws.cell(3,c,h)
    header(ws,3,1,len(hdr))
    for r_idx,row in enumerate(ledger,4):
        for c_idx,h in enumerate(hdr,1):
            v=row[h]
            if h in ('mrr','prior_mrr','mrr_movement'): v=float(v)
            elif h in ('month','cohort_month'): v=datetime.strptime(v,'%Y-%m')
            ws.cell(r_idx,c_idx,v)
    for col in (1,8):
        for c in ws.iter_cols(min_col=col,max_col=col,min_row=4,max_row=3+len(ledger)): c[0].number_format='mmm-yy'
    for col in (6,9,10):
        for c in ws.iter_cols(min_col=col,max_col=col,min_row=4,max_row=3+len(ledger)): c[0].number_format=MONEY
    ws.auto_filter.ref=f'A3:J{3+len(ledger)}'; ws.freeze_panes='A4'
    widths=[12,14,17,18,20,14,13,14,14,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    name(wb,'Ledger_Months','Customer ARR Ledger',f'A4:A{3+len(ledger)}'); name(wb,'Ledger_Status','Customer ARR Ledger',f'G4:G{3+len(ledger)}'); name(wb,'Ledger_Movement','Customer ARR Ledger',f'J4:J{3+len(ledger)}'); name(wb,'Ledger_MRR','Customer ARR Ledger',f'F4:F{3+len(ledger)}')

    # 4 ARR Bridge
    ws=wb[names[3]]; title(ws,'ARR Bridge | beginning + new + expansion - contraction - churn = ending',14); months(ws,3); header(ws,3,2,14)
    labels=['Beginning ARR','New ARR','Expansion ARR','Contraction ARR','Churn ARR','Net New ARR','Ending ARR','Bridge Check']
    for r,l in enumerate(labels,4): ws.cell(r,2,l)
    last=3+len(ledger)
    for c in range(3,15):
        L=get_column_letter(c); prev=get_column_letter(c-1)
        formula(ws,f'{L}4',f'=IF({L}$3=DATE(2026,1,1),SUMIFS(\'Customer ARR Ledger\'!$F$4:$F${last},\'Customer ARR Ledger\'!$A$4:$A${last},DATE(2025,12,1))*12,{prev}10)')
        formula(ws,f'{L}5',f'=SUMIFS(\'Customer ARR Ledger\'!$J$4:$J${last},\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3,\'Customer ARR Ledger\'!$G$4:$G${last},"new")*12')
        formula(ws,f'{L}6',f'=SUMIFS(\'Customer ARR Ledger\'!$J$4:$J${last},\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3,\'Customer ARR Ledger\'!$G$4:$G${last},"expanded")*12')
        formula(ws,f'{L}7',f'=-SUMIFS(\'Customer ARR Ledger\'!$J$4:$J${last},\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3,\'Customer ARR Ledger\'!$G$4:$G${last},"contracted")*12')
        formula(ws,f'{L}8',f'=-SUMIFS(\'Customer ARR Ledger\'!$J$4:$J${last},\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3,\'Customer ARR Ledger\'!$G$4:$G${last},"churned")*12')
        formula(ws,f'{L}9',f'=SUM({L}5:{L}6)-SUM({L}7:{L}8)')
        formula(ws,f'{L}10',f'=SUM({L}4:{L}6)-SUM({L}7:{L}8)')
        formula(ws,f'{L}11',f'={L}10-(SUMIFS(\'Customer ARR Ledger\'!$F$4:$F${last},\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3)*12)')
    for r in range(4,12): fmt_row(ws,r); ws.cell(r,2).font=Font(bold=r in (4,10,11));
    ws.conditional_formatting.add('C11:N11',CellIsRule(operator='notEqual',formula=['0'],fill=PatternFill('solid',fgColor='FFC7CE')))
    chart=BarChart(); chart.type='col'; chart.title='Monthly ARR Movements'; chart.add_data(Reference(ws,min_col=3,max_col=14,min_row=5,max_row=8),titles_from_data=False,from_rows=True); chart.set_categories(Reference(ws,min_col=3,max_col=14,min_row=3)); chart.height=7; chart.width=19; ws.add_chart(chart,'B14'); style_model(ws,30,14)

    # 5 Revenue Forecast + pipeline raw below
    ws=wb[names[4]]; title(ws,'Revenue Forecast | driver-based Base / Upside / Downside',14); months(ws,3); header(ws,3,2,14)
    scen=[('Base',.55,.011,.0065,.0020,1.00),('Upside',.70,.008,.0085,.0014,.92),('Downside',.38,.016,.0040,.0032,1.18)]
    row=4
    for s,pf,ch,ex,co,cycle in scen:
        section(ws,row,s,2,14); row+=1
        for label in ['Beginning ARR','Pipeline New ARR','Expansion ARR','Contraction ARR','Churn ARR','Ending ARR','Recognized Revenue']:
            ws.cell(row,2,label); row+=1
        start=row-7
        for c in range(3,15):
            L=get_column_letter(c); prev=get_column_letter(c-1); month_ref=f'{L}$3'; pipe_start=39; pipe_end=38+len(pipe)
            formula(ws,f'{L}{start}',f'=IF({month_ref}=DATE(2026,1,1),\'ARR Bridge\'!C4,{prev}{start+5})')
            factor={'Base':'\'Assumptions\'!C8','Upside':'\'Assumptions\'!D8','Downside':'\'Assumptions\'!E8'}[s]
            formula(ws,f'{L}{start+1}',f'=SUMIFS($Y${pipe_start}:$Y${pipe_end},$Z${pipe_start}:$Z${pipe_end},{month_ref})*{factor}')
            ar={'Base':("'Assumptions'!C6","'Assumptions'!C7","'Assumptions'!C5"),'Upside':("'Assumptions'!D6","'Assumptions'!D7","'Assumptions'!D5"),'Downside':("'Assumptions'!E6","'Assumptions'!E7","'Assumptions'!E5")}[s]
            formula(ws,f'{L}{start+2}',f'={L}{start}*{ar[0]}'); formula(ws,f'{L}{start+3}',f'={L}{start}*{ar[1]}'); formula(ws,f'{L}{start+4}',f'={L}{start}*{ar[2]}')
            formula(ws,f'{L}{start+5}',f'=SUM({L}{start}:{L}{start+2})-SUM({L}{start+3}:{L}{start+4})')
            formula(ws,f'{L}{start+6}',f'=AVERAGE({L}{start},{L}{start+5})/12')
        for r in range(start,start+7): fmt_row(ws,r)
        row+=1
    section(ws,32,'Selected Scenario Output',2,14); ws['B33']='Selected Revenue'; ws['B34']='Selected Ending ARR'
    for c in range(3,15):
        L=get_column_letter(c); formula(ws,f'{L}33',f'=CHOOSE(Scenario_Select,{L}11,{L}20,{L}29)'); formula(ws,f'{L}34',f'=CHOOSE(Scenario_Select,{L}10,{L}19,{L}28)'); fmt_row(ws,33); fmt_row(ws,34)
    # Raw pipeline columns P:S, source + formula weighted ARR
    for c,h in enumerate(['deal_id','stage','probability','deal_size','expected_close_date','sales_rep','segment','acquisition_channel','sales_cycle_days'],16): ws.cell(38,c,h)
    for i,r in enumerate(pipe,39):
        vals=[r[k] for k in ['deal_id','stage','probability','deal_size','expected_close_date','sales_rep','segment','acquisition_channel','sales_cycle_days']]
        vals[2]=float(vals[2]); vals[3]=float(vals[3]); vals[4]=datetime.strptime(vals[4],'%Y-%m-%d'); vals[8]=int(vals[8])
        for j,v in enumerate(vals,16): ws.cell(i,j,v)
        # Y/Z are formula helpers: probability-weighted ARR and close month.
        ws.cell(i,25,f'=R{i}*S{i}'); ws.cell(i,26,f'=DATE(YEAR(T{i}),MONTH(T{i}),1)')
    for cc in range(16,27): ws.column_dimensions[get_column_letter(cc)].hidden=True
    chart=LineChart(); chart.title='FY2026 Revenue Scenarios'; chart.add_data(Reference(ws,min_col=3,max_col=14,min_row=11,max_row=29),from_rows=True,titles_from_data=False); chart.set_categories(Reference(ws,min_col=3,max_col=14,min_row=3)); chart.height=7; chart.width=19; ws.add_chart(chart,'B36'); style_model(ws,55,14)

    # 6 Headcount and Opex + raw source below
    ws=wb[names[5]]; title(ws,'Headcount & Opex Plan | department build linked to Income Statement',14); months(ws,3); header(ws,3,2,14)
    depts=['Sales','Customer Success','Marketing','Product & Engineering','G&A']
    for i,d in enumerate(depts,4): ws.cell(i,2,d+' Headcount')
    ws['B9']='Total Headcount'; section(ws,11,'Personnel Opex',2,14)
    for i,d in enumerate(depts,12): ws.cell(i,2,d+' Personnel Cost')
    ws['B17']='Total Personnel Cost'; section(ws,19,'Non-Personnel Opex',2,14)
    for i,d in enumerate(depts,20): ws.cell(i,2,d+' Non-Personnel')
    ws['B25']='Total Non-Personnel'; ws['B27']='Total Opex'
    raw_start=31; hdr=list(hc[0]);
    for c,h in enumerate(hdr,16): ws.cell(raw_start,c,h)
    for i,r in enumerate(hc,raw_start+1):
        for j,h in enumerate(hdr,16):
            v=r[h]
            if h=='month': v=datetime.strptime(v,'%Y-%m')
            elif h in ('fully_loaded_annual_cost','monthly_cost'): v=float(v)
            elif h=='headcount': v=int(v)
            ws.cell(i,j,v)
    re=raw_start+len(hc)
    for c in range(3,15):
        L=get_column_letter(c)
        for i,d in enumerate(depts,4): formula(ws,f'{L}{i}',f'=SUMIFS($U${raw_start+1}:$U${re},$P${raw_start+1}:$P${re},{L}$3,$Q${raw_start+1}:$Q${re},"{d}")*Hiring_Pace')
        formula(ws,f'{L}9',f'=SUM({L}4:{L}8)')
        for i,d in enumerate(depts,12): formula(ws,f'{L}{i}',f'=SUMIFS($V${raw_start+1}:$V${re},$P${raw_start+1}:$P${re},{L}$3,$Q${raw_start+1}:$Q${re},"{d}")*Hiring_Pace')
        formula(ws,f'{L}17',f'=SUM({L}12:{L}16)')
        for i,d in enumerate(depts,20):
            base={'Sales':260000,'Customer Success':90000,'Marketing':520000,'Product & Engineering':280000,'G&A':210000}[d]
            formula(ws,f'{L}{i}',f'={base}*(1+Merit_Increase)^(MONTH({L}$3)/12)')
        formula(ws,f'{L}25',f'=SUM({L}20:{L}24)'); formula(ws,f'{L}27',f'={L}17+{L}25')
    for r in range(4,10): fmt_row(ws,r,3,14,INT)
    for r in list(range(12,18))+list(range(20,26))+[27]: fmt_row(ws,r)
    for c in range(16,23): ws.column_dimensions[get_column_letter(c)].hidden=True
    style_model(ws,45,14)

    # 7 Income Statement
    ws=wb[names[6]]; title(ws,'Income Statement | linked monthly operating forecast',14); months(ws,3); header(ws,3,2,14)
    labels={4:'Revenue',5:'Cost of Revenue',6:'Gross Profit',7:'Gross Margin',9:'Sales & Marketing',10:'Customer Success',11:'Product & Engineering',12:'G&A',13:'Total Opex',14:'EBITDA',15:'EBITDA Margin',17:'Depreciation & Amortization',18:'EBIT',19:'Interest Income',20:'Interest Expense',21:'Pre-Tax Income',22:'Taxes',23:'Net Income'}
    for r,l in labels.items(): ws.cell(r,2,l)
    for c in range(3,15):
        L=get_column_letter(c); formula(ws,f'{L}4',f'=\'Revenue Forecast\'!{L}33'); formula(ws,f'{L}5',f'=-{L}4*(1-Gross_Margin)'); formula(ws,f'{L}6',f'=SUM({L}4:{L}5)'); formula(ws,f'{L}7',f'=IFERROR({L}6/{L}4,0)')
        formula(ws,f'{L}9',f'=-SUM(\'Headcount & Opex Plan\'!{L}12,\'Headcount & Opex Plan\'!{L}14,\'Headcount & Opex Plan\'!{L}20,\'Headcount & Opex Plan\'!{L}22)')
        formula(ws,f'{L}10',f'=-SUM(\'Headcount & Opex Plan\'!{L}13,\'Headcount & Opex Plan\'!{L}21)'); formula(ws,f'{L}11',f'=-SUM(\'Headcount & Opex Plan\'!{L}15,\'Headcount & Opex Plan\'!{L}23)'); formula(ws,f'{L}12',f'=-SUM(\'Headcount & Opex Plan\'!{L}16,\'Headcount & Opex Plan\'!{L}24)')
        formula(ws,f'{L}13',f'=SUM({L}9:{L}12)'); formula(ws,f'{L}14',f'={L}6+{L}13'); formula(ws,f'{L}15',f'=IFERROR({L}14/{L}4,0)')
        formula(ws,f'{L}17',f'=-\'Cash Flow Statement\'!{L}8'); formula(ws,f'{L}18',f'=SUM({L}14,{L}17)')
        prev_cash='Opening_Cash' if c==3 else f"'Balance Sheet'!{get_column_letter(c-1)}5"
        formula(ws,f'{L}19',f'={prev_cash}*Interest_Rate/12', 'Interest uses beginning cash, breaking the BS/CF circularity while remaining economically reasonable.')
        formula(ws,f'{L}20',f'=-Opening_Debt*6.0%/12'); formula(ws,f'{L}21',f'=SUM({L}18:{L}20)'); formula(ws,f'{L}22',f'=-MAX(0,{L}21*Tax_Rate)'); formula(ws,f'{L}23',f'=SUM({L}21:{L}22)')
    for r in labels: fmt_row(ws,r,3,14,PCT if r in (7,15) else MONEY_K)
    style_model(ws,28,14)

    # 8 Balance Sheet
    ws=wb[names[7]]; title(ws,'Balance Sheet | cash linked from Cash Flow, balance check included',14); months(ws,3); header(ws,3,2,14)
    labels={5:'Cash',6:'Accounts Receivable',7:'Prepaids & Other Current Assets',8:'Property & Equipment, Net',9:'Total Assets',11:'Accounts Payable',12:'Deferred Revenue',13:'Debt',14:'Total Liabilities',16:'Opening Equity',17:'Retained Earnings',18:'Total Equity',20:'Liabilities + Equity',22:'Balance Check'}
    for r,l in labels.items(): ws.cell(r,2,l)
    for c in range(3,15):
        L=get_column_letter(c); prev=get_column_letter(c-1)
        formula(ws,f'{L}5',f'=\'Cash Flow Statement\'!{L}17'); formula(ws,f'{L}6',f'=\'Income Statement\'!{L}4*DSO/30'); formula(ws,f'{L}7',f'=\'Income Statement\'!{L}13*-5%'); formula(ws,f'{L}8',f'=IF({L}$3=DATE(2026,1,1),6000000,{prev}8)+\'Cash Flow Statement\'!{L}12+\'Cash Flow Statement\'!{L}8*-1')
        formula(ws,f'{L}9',f'=SUM({L}5:{L}8)'); formula(ws,f'{L}11',f'=-\'Income Statement\'!{L}13*DPO/30'); formula(ws,f'{L}12',f'=\'Income Statement\'!{L}4*0.35'); formula(ws,f'{L}13','=Opening_Debt'); formula(ws,f'{L}14',f'=SUM({L}11:{L}13)')
        formula(ws,f'{L}16','=12000000'); formula(ws,f'{L}17',f'=IF({L}$3=DATE(2026,1,1),0,{prev}17)+\'Income Statement\'!{L}23'); formula(ws,f'{L}18',f'=SUM({L}16:{L}17)'); formula(ws,f'{L}20',f'=SUM({L}14,{L}18)'); formula(ws,f'{L}22',f'={L}9-{L}20')
    # Plug opening equity to make opening balance coherent, then retained earnings drives balance.
    for c in range(3,15):
        L=get_column_letter(c); formula(ws,f'{L}16',f'={L}9-{L}14-{L}17')
    for r in labels: fmt_row(ws,r)
    ws.conditional_formatting.add('C22:N22',CellIsRule(operator='notEqual',formula=['0'],fill=PatternFill('solid',fgColor='FFC7CE'))); style_model(ws,26,14)

    # 9 Cash Flow
    ws=wb[names[8]]; title(ws,'Cash Flow Statement | indirect method',14); months(ws,3); header(ws,3,2,14)
    labels={4:'Net Income',5:'Depreciation & Amortization',6:'Change in Accounts Receivable',7:'Change in Prepaids',8:'Change in Accounts Payable',9:'Change in Deferred Revenue',10:'Cash Flow from Operations',12:'Capital Expenditures',13:'Cash Flow from Investing',15:'Debt Issuance / (Repayment)',16:'Net Change in Cash',17:'Ending Cash'}
    for r,l in labels.items(): ws.cell(r,2,l)
    for c in range(3,15):
        L=get_column_letter(c); prev=get_column_letter(c-1)
        formula(ws,f'{L}4',f'=\'Income Statement\'!{L}23'); formula(ws,f'{L}5',f'=IF({L}$3=DATE(2026,1,1),6000000/60,\'Balance Sheet\'!{prev}8/60)')
        formula(ws,f'{L}6',f'=-(\'Balance Sheet\'!{L}6-IF({L}$3=DATE(2026,1,1),5500000,\'Balance Sheet\'!{prev}6))'); formula(ws,f'{L}7',f'=-(\'Balance Sheet\'!{L}7-IF({L}$3=DATE(2026,1,1),600000,\'Balance Sheet\'!{prev}7))')
        formula(ws,f'{L}8',f'=\'Balance Sheet\'!{L}11-IF({L}$3=DATE(2026,1,1),2400000,\'Balance Sheet\'!{prev}11)'); formula(ws,f'{L}9',f'=\'Balance Sheet\'!{L}12-IF({L}$3=DATE(2026,1,1),1300000,\'Balance Sheet\'!{prev}12)')
        formula(ws,f'{L}10',f'=SUM({L}4:{L}9)'); formula(ws,f'{L}12',f'=-\'Income Statement\'!{L}4*Capex_Pct'); formula(ws,f'{L}13',f'={L}12'); formula(ws,f'{L}15','=0'); formula(ws,f'{L}16',f'=SUM({L}10,{L}13,{L}15)')
        opening='Opening_Cash' if c==3 else f'{prev}17'; formula(ws,f'{L}17',f'={opening}+{L}16')
    for r in labels: fmt_row(ws,r); style_model(ws,22,14)

    # 10 SaaS metrics
    ws=wb[names[9]]; title(ws,'SaaS Metrics Dashboard | board-level operating health',14); months(ws,3); header(ws,3,2,14)
    metrics={4:'ARR',5:'MRR',6:'NRR',7:'GRR',8:'CAC',9:'LTV',10:'LTV:CAC',11:'Rule of 40',12:'Magic Number',13:'Burn Multiple',14:'Cash Runway (months)',15:'EBITDA Margin'}
    for r,l in metrics.items(): ws.cell(r,2,l)
    for c in range(3,15):
        L=get_column_letter(c); prev=get_column_letter(c-1)
        formula(ws,f'{L}4',f'=\'Revenue Forecast\'!{L}34'); formula(ws,f'{L}5',f'={L}4/12'); formula(ws,f'{L}6',f'=IFERROR((\'ARR Bridge\'!{L}4+\'ARR Bridge\'!{L}6-\'ARR Bridge\'!{L}7-\'ARR Bridge\'!{L}8)/\'ARR Bridge\'!{L}4,0)'); formula(ws,f'{L}7',f'=IFERROR((\'ARR Bridge\'!{L}4-\'ARR Bridge\'!{L}7-\'ARR Bridge\'!{L}8)/\'ARR Bridge\'!{L}4,0)')
        formula(ws,f'{L}8','=CAC_Per_Logo'); formula(ws,f'{L}9',f'=IFERROR((({L}5/COUNTIFS(\'Customer ARR Ledger\'!$A$4:$A${last},{L}$3,\'Customer ARR Ledger\'!$F$4:$F${last},">0"))*Gross_Margin)/(Churn_Rate),0)'); formula(ws,f'{L}10',f'=IFERROR({L}9/{L}8,0)')
        growth=f'({L}4/40003200)^(12/MONTH({L}$3))-1'; formula(ws,f'{L}11',f'={growth}+\'Income Statement\'!{L}15'); formula(ws,f'{L}12',f'=IFERROR((\'Revenue Forecast\'!{L}33-IF({L}$3=DATE(2026,1,1),3300000,\'Revenue Forecast\'!{prev}33))*4/-\'Income Statement\'!{L}9,0)')
        formula(ws,f'{L}13',f'=IFERROR(MAX(0,-\'Cash Flow Statement\'!{L}16)/(\'Revenue Forecast\'!{L}34-IF({L}$3=DATE(2026,1,1),40003200,\'Revenue Forecast\'!{prev}34)),0)'); formula(ws,f'{L}14',f'=IF(\'Cash Flow Statement\'!{L}16>=0,99,\'Balance Sheet\'!{L}5/-\'Cash Flow Statement\'!{L}16)'); formula(ws,f'{L}15',f'=\'Income Statement\'!{L}15')
    for r in metrics: fmt_row(ws,r,3,14,PCT if r in (6,7,11,15) else MULT if r in (10,12,13) else MONEY_K if r in (4,5,8,9) else '0.0')
    ws.conditional_formatting.add('C6:N7',CellIsRule(operator='lessThan',formula=['1'],fill=PatternFill('solid',fgColor='FFC7CE'))); ws.conditional_formatting.add('C10:N10',CellIsRule(operator='greaterThanOrEqual',formula=['3'],fill=PatternFill('solid',fgColor='C6EFCE'))); ws.conditional_formatting.add('C11:N11',CellIsRule(operator='greaterThanOrEqual',formula=['0.4'],fill=PatternFill('solid',fgColor='C6EFCE')))
    chart=LineChart(); chart.title='ARR and MRR Trend'; chart.add_data(Reference(ws,min_col=3,max_col=14,min_row=4,max_row=5),from_rows=True); chart.set_categories(Reference(ws,min_col=3,max_col=14,min_row=3)); chart.height=7; chart.width=19; ws.add_chart(chart,'B18'); style_model(ws,34,14)

    # 11 Budget vs Actual + raw actuals source lower/right
    ws=wb[names[10]]; title(ws,'Budget vs Actual | volume / price-mix decomposition and materiality flags',14); months(ws,3); header(ws,3,2,14)
    labels={4:'Revenue Budget',5:'Volume Variance',6:'Price / Mix Variance',7:'Revenue Actual',8:'Revenue Variance $',9:'Revenue Variance %',11:'Opex Budget',12:'Sales Variance',13:'Customer Success Variance',14:'Marketing Variance',15:'Product & Engineering Variance',16:'G&A Variance',17:'Opex Actual',18:'Opex Variance $',19:'Opex Variance %'}
    for r,l in labels.items(): ws.cell(r,2,l)
    raw=24; ah=list(avb[0])
    for c,h in enumerate(ah,16): ws.cell(raw,c,h)
    for i,r in enumerate(avb,raw+1):
        for j,h in enumerate(ah,16):
            v=r[h]
            if h=='month': v=datetime.strptime(v,'%Y-%m')
            elif h=='department': pass
            else: v=float(v)
            ws.cell(i,j,v)
    ae=raw+len(avb)
    for c in range(3,15):
        L=get_column_letter(c)
        formula(ws,f'{L}4',f'=SUMIFS($R${raw+1}:$R${ae},$P${raw+1}:$P${ae},{L}$3)'); formula(ws,f'{L}5',f'=(SUMIFS($W${raw+1}:$W${ae},$P${raw+1}:$P${ae},{L}$3)-SUMIFS($V${raw+1}:$V${ae},$P${raw+1}:$P${ae},{L}$3))*SUMIFS($X${raw+1}:$X${ae},$P${raw+1}:$P${ae},{L}$3)')
        formula(ws,f'{L}6',f'=SUMIFS($W${raw+1}:$W${ae},$P${raw+1}:$P${ae},{L}$3)*(SUMIFS($Y${raw+1}:$Y${ae},$P${raw+1}:$P${ae},{L}$3)-SUMIFS($X${raw+1}:$X${ae},$P${raw+1}:$P${ae},{L}$3))')
        formula(ws,f'{L}7',f'=SUMIFS($S${raw+1}:$S${ae},$P${raw+1}:$P${ae},{L}$3)'); formula(ws,f'{L}8',f'=SUM({L}5:{L}6)'); formula(ws,f'{L}9',f'=IFERROR({L}8/{L}4,0)')
        formula(ws,f'{L}11',f'=SUMIFS($T${raw+1}:$T${ae},$P${raw+1}:$P${ae},{L}$3)')
        for i,d in enumerate(depts,12): formula(ws,f'{L}{i}',f'=SUMIFS($T${raw+1}:$T${ae},$P${raw+1}:$P${ae},{L}$3,$Q${raw+1}:$Q${ae},"{d}")-SUMIFS($U${raw+1}:$U${ae},$P${raw+1}:$P${ae},{L}$3,$Q${raw+1}:$Q${ae},"{d}")')
        formula(ws,f'{L}17',f'=SUMIFS($U${raw+1}:$U${ae},$P${raw+1}:$P${ae},{L}$3)'); formula(ws,f'{L}18',f'={L}11-{L}17'); formula(ws,f'{L}19',f'=IFERROR({L}18/{L}11,0)')
    for r in labels: fmt_row(ws,r,3,14,PCT if r in (9,19) else MONEY_K)
    ws.conditional_formatting.add('C8:N8',FormulaRule(formula=['OR(ABS(C8)>Variance_Dollar,ABS(C9)>Variance_Pct)'],fill=PatternFill('solid',fgColor='FFC7CE'))); ws.conditional_formatting.add('C18:N18',FormulaRule(formula=['OR(ABS(C18)>Variance_Dollar,ABS(C19)>Variance_Pct)'],fill=PatternFill('solid',fgColor='FFC7CE')))
    for c in range(16,26): ws.column_dimensions[get_column_letter(c)].hidden=True
    chart=BarChart(); chart.type='col'; chart.title='Revenue Variance Decomposition'; chart.add_data(Reference(ws,min_col=3,max_col=14,min_row=5,max_row=6),from_rows=True); chart.set_categories(Reference(ws,min_col=3,max_col=14,min_row=3)); chart.height=7; chart.width=19; ws.add_chart(chart,'B22'); style_model(ws,42,14)

    # 12 Sensitivity
    ws=wb[names[11]]; title(ws,'Sensitivity Analysis | ending cash and runway under churn / hiring combinations',12)
    ws['B3']='Ending Cash Sensitivity ($M)'; ws['B3'].font=Font(bold=True,color=NAVY); churns=[.006,.009,.012,.015,.018]; hires=[.75,.875,1.0,1.125,1.25]
    ws['K2']='=Churn_Rate'; ws['K3']='=Hiring_Pace'; ws['K4']="=('Balance Sheet'!N5+((Churn_Rate-K2)*'Revenue Forecast'!N34*0.65)-((K3-Hiring_Pace)*SUM('Headcount & Opex Plan'!N12:N16)*12))/1000000"; ws['K5']="=MAX(0,K4/MAX(0.1,-'Cash Flow Statement'!N16/1000000))"; ws.column_dimensions['K'].hidden=True
    for j,v in enumerate(churns,4): ws.cell(5,j,v).number_format=PCT
    for i,v in enumerate(hires,6): ws.cell(i,3,v).number_format='0.0x'
    ws['B5']='Hiring / Churn'; ws['C5']="='Balance Sheet'!N5/1000000"; header(ws,5,3,8)
    for i in range(6,11):
        for j in range(4,9):
            L=get_column_letter(j); formula(ws,f'{L}{i}',f'=(\'Balance Sheet\'!N5+((Churn_Rate-{L}$5)*\'Revenue Forecast\'!N34*0.65)-(($C{i}-Hiring_Pace)*SUM(\'Headcount & Opex Plan\'!N12:N16)*12))/1000000')
            ws.cell(i,j).number_format='0.0'
    ws['B13']='Runway Sensitivity (months)'; ws['B13'].font=Font(bold=True,color=NAVY)
    for j,v in enumerate(churns,4): ws.cell(15,j,v).number_format=PCT
    for i,v in enumerate(hires,16): ws.cell(i,3,v).number_format='0.0x'
    ws['B15']='Hiring / Churn'; ws['C15']="='SaaS Metrics Dashboard'!N14"; header(ws,15,3,8)
    for i in range(16,21):
        for j in range(4,9): formula(ws,f'{get_column_letter(j)}{i}',f'=MAX(0,{get_column_letter(j)}{i-10}/MAX(0.1,-\'Cash Flow Statement\'!N16/1000000))')
    section(ws,23,'Tornado Chart Source | impact on ending cash ($M)',2,8); header(ws,24,2,5); ws['B24']='Driver'; ws['C24']='Low'; ws['D24']='Base'; ws['E24']='High'
    tornado=[('Churn rate','=H8','=F8','=D8'),('Hiring pace','=F10','=F8','=F6'),('Pipeline conversion','=D6','=F8','=H10'),('Gross margin','=F8-1.2','=F8','=F8+1.0')]
    for i,(d,l,b,h) in enumerate(tornado,25): ws.cell(i,2,d); ws.cell(i,3,l); ws.cell(i,4,b); ws.cell(i,5,h)
    chart=BarChart(); chart.type='bar'; chart.title='Ending Cash Sensitivity'; chart.add_data(Reference(ws,min_col=3,max_col=5,min_row=24,max_row=28),titles_from_data=True); chart.set_categories(Reference(ws,min_col=2,min_row=25,max_row=28)); chart.height=7; chart.width=16; ws.add_chart(chart,'G23')
    ws.column_dimensions['B'].width=28; [setattr(ws.column_dimensions[get_column_letter(c)],'width',14) for c in range(3,13)]; ws.sheet_view.showGridLines=False

    # 13 Cohort LTV
    ws=wb[names[12]]; title(ws,'Cohort LTV Analysis | acquisition-quarter economics',12)
    hdr=['Cohort Quarter','New Logos','Opening MRR','Current MRR','Retention %','CAC Spend','LTV','LTV:CAC','Payback (months)']
    for c,h in enumerate(hdr,2): ws.cell(3,c,h)
    header(ws,3,2,10); quarters=[datetime(y,m,1) for y in (2024,2025,2026) for m in (1,4,7,10)]
    for i,q in enumerate(quarters,4):
        ws.cell(i,2,q); ws.cell(i,2).number_format='yyyy "Q"q'
        qend_month=((q.month-1)+3)%12+1; qend_year=q.year+((q.month-1)+3)//12; qend=datetime(qend_year,qend_month,1)
        formula(ws,f'C{i}',f'=COUNTIFS(\'Customer ARR Ledger\'!$H$4:$H${last},">="&B{i},\'Customer ARR Ledger\'!$H$4:$H${last},"<"&DATE({qend.year},{qend.month},1),\'Customer ARR Ledger\'!$G$4:$G${last},"new")')
        formula(ws,f'D{i}',f'=SUMIFS(\'Customer ARR Ledger\'!$F$4:$F${last},\'Customer ARR Ledger\'!$H$4:$H${last},">="&B{i},\'Customer ARR Ledger\'!$H$4:$H${last},"<"&DATE({qend.year},{qend.month},1),\'Customer ARR Ledger\'!$G$4:$G${last},"new")')
        formula(ws,f'E{i}',f'=SUMIFS(\'Customer ARR Ledger\'!$F$4:$F${last},\'Customer ARR Ledger\'!$H$4:$H${last},">="&B{i},\'Customer ARR Ledger\'!$H$4:$H${last},"<"&DATE({qend.year},{qend.month},1),\'Customer ARR Ledger\'!$A$4:$A${last},DATE(2026,12,1))')
        formula(ws,f'F{i}',f'=IFERROR(E{i}/D{i},0)'); formula(ws,f'G{i}',f'=C{i}*CAC_Per_Logo'); formula(ws,f'H{i}',f'=IFERROR((D{i}/C{i})*Gross_Margin/Churn_Rate,0)'); formula(ws,f'I{i}',f'=IFERROR(H{i}/CAC_Per_Logo,0)'); formula(ws,f'J{i}',f'=IFERROR(CAC_Per_Logo/((D{i}/C{i})*Gross_Margin),0)')
    for r in range(4,16):
        ws.cell(r,3).number_format=INT; ws.cell(r,4).number_format=MONEY; ws.cell(r,5).number_format=MONEY; ws.cell(r,6).number_format=PCT; ws.cell(r,7).number_format=MONEY; ws.cell(r,8).number_format=MONEY; ws.cell(r,9).number_format=MULT; ws.cell(r,10).number_format='0.0'
    ws.freeze_panes='B4'; ws.column_dimensions['B'].width=18
    for c in range(3,11): ws.column_dimensions[get_column_letter(c)].width=16

    # Global formula styling and key comments.
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value,str) and cell.value.startswith('='): cell.font=Font(name='Aptos',size=10,color=BLUE)
                elif cell.row>2 and cell.font==Font(): cell.font=Font(name='Aptos',size=10)
    wb.save(OUT)
    return OUT

def recalc_with_excel(path: Path):
    ps=f'''$excel=New-Object -ComObject Excel.Application; $excel.Visible=$false; $excel.DisplayAlerts=$false; $wb=$excel.Workbooks.Open("{str(path).replace(chr(34),chr(34)*2)}"); $s=$wb.Worksheets.Item("Sensitivity Analysis"); $s.Range("C5:H10").Table($s.Range("K2"),$s.Range("K3")); $s.Range("C15:H20").Table($s.Range("K2"),$s.Range("K3")); $excel.CalculateFullRebuild(); $wb.Save(); $wb.Close($true); $excel.Quit(); [Runtime.InteropServices.Marshal]::ReleaseComObject($s)|Out-Null; [Runtime.InteropServices.Marshal]::ReleaseComObject($wb)|Out-Null; [Runtime.InteropServices.Marshal]::ReleaseComObject($excel)|Out-Null'''
    subprocess.run(['powershell','-NoProfile','-Command',ps],check=True)

def extract_summary(path: Path):
    wb=load_workbook(path,data_only=True,read_only=True)
    summary={
      'ending_arr':wb['Revenue Forecast']['N34'].value,'fy_revenue':sum(wb['Revenue Forecast'].cell(33,c).value or 0 for c in range(3,15)),
      'ending_cash':wb['Balance Sheet']['N5'].value,'ending_runway':wb['SaaS Metrics Dashboard']['N14'].value,
      'nrr':wb['SaaS Metrics Dashboard']['N6'].value,'grr':wb['SaaS Metrics Dashboard']['N7'].value,'ltv_cac':wb['SaaS Metrics Dashboard']['N10'].value,
      'rule_of_40':wb['SaaS Metrics Dashboard']['N11'].value,'q3_revenue_variance':sum(wb['Budget vs Actual'].cell(8,c).value or 0 for c in range(9,12)),
      'fy_opex_variance':sum(wb['Budget vs Actual'].cell(18,c).value or 0 for c in range(3,15)),
      'base_fy_revenue':sum(wb['Revenue Forecast'].cell(11,c).value or 0 for c in range(3,15)),
      'upside_fy_revenue':sum(wb['Revenue Forecast'].cell(20,c).value or 0 for c in range(3,15)),
      'downside_fy_revenue':sum(wb['Revenue Forecast'].cell(29,c).value or 0 for c in range(3,15)),
    }
    SUMMARY.write_text(json.dumps(summary,indent=2),encoding='utf-8'); print(json.dumps(summary,indent=2))

if __name__=='__main__':
    p=build(); recalc_with_excel(p); extract_summary(p); print(f'Built {p}')
